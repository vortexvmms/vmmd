#!/usr/bin/env python3
"""
VMMS nightly report emailer.

Builds two reports for the current month and emails them (PDF + Excel)
to a Gmail inbox via Gmail SMTP. Designed to run from GitHub Actions on a
nightly cron. No secrets are hard-coded; everything sensitive comes from
environment variables (GitHub Actions secrets).

Env vars required (set as GitHub Actions secrets):
  BOT_EMAIL            a VMMS login e-mail with full access (reads all sites)
  BOT_PASSWORD         that login's password
  GMAIL_USER           the Gmail address that sends the mail
  GMAIL_APP_PASSWORD   16-char Google App Password for that Gmail
  REPORT_TO            where to send the report (usually the same Gmail)

Optional:
  REPORT_TEST=1        use built-in sample data, build files, DO NOT send
"""

import os
import sys
import smtplib
import datetime as dt
from email.message import EmailMessage

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF
from fpdf.enums import XPos, YPos

# ---- public (non-secret) config, mirrors frontend/js/config.js ----
SUPABASE_URL = "https://lqnbdemtgkermhaqfboh.supabase.co"
SUPABASE_PUBLISHABLE = "sb_publishable_GV2oQS2wP2ltMeg2Teh1Rw_qT3lLGno"
BACKEND_URL = "https://vmms-backend-7j1v.onrender.com"
SGT = dt.timezone(dt.timedelta(hours=8))

# brand colours
RED = "C00000"
GREY = "595959"
LIGHT = "F2F2F2"


def log(*a):
    print(*a, flush=True)


def envv(name, default=None):
    """Read an env var and strip stray spaces / newlines (secrets often
    get pasted with a trailing line break, which breaks e-mail headers)."""
    v = os.environ.get(name, default)
    return v.strip() if isinstance(v, str) else v


def latin(x):
    """fpdf core fonts are latin-1; make any text safe."""
    return str("" if x is None else x).encode("latin-1", "replace").decode("latin-1")


# ---------------------------------------------------------------- data
def get_token():
    r = requests.post(
        f"{SUPABASE_URL}/auth/v1/token?grant_type=password",
        headers={"apikey": SUPABASE_PUBLISHABLE, "Content-Type": "application/json"},
        json={"email": envv("BOT_EMAIL"), "password": envv("BOT_PASSWORD")},
        timeout=30,
    )
    if r.status_code != 200:
        raise SystemExit(f"Login failed ({r.status_code}): {r.text[:200]}")
    return r.json()["access_token"]


def fetch_attendance(token, dfrom, dto):
    r = requests.get(
        f"{BACKEND_URL}/api/v1/reports/attendance",
        params={"dfrom": dfrom, "dto": dto},
        headers={"Authorization": f"Bearer {token}"},
        timeout=120,
    )
    if r.status_code != 200:
        raise SystemExit(f"Attendance fetch failed ({r.status_code}): {r.text[:200]}")
    return r.json()


def sample_rows(today):
    """Test data so the build path can be exercised without network."""
    mk = lambda d, code, name, site, pres, s, e, nh, ot, sub, absc="": {
        "date": d, "site": site, "worker_code": code, "worker": name,
        "present": pres, "absence": absc, "start": s, "end": e,
        "nh": nh, "ot": ot, "day_type": "normal", "submitted": sub,
    }
    d0 = today.replace(day=1)
    rows = []
    for i in range(0, (today - d0).days + 1):
        d = (d0 + dt.timedelta(days=i)).isoformat()
        rows.append(mk(d, "W001", "Rajan Kumar", "Tuas Depot", True, "08:00", "17:00", 8, 1.5, True))
        rows.append(mk(d, "W002", "Muthu Samy", "Changi P1", True, "08:00", "18:00", 8, 2, True))
        rows.append(mk(d, "W003", "Ah Meng", "Tuas Depot", False, "", "", 0, 0, False, "mc"))
    # today: leave one unverified to show the pending highlight
    td = today.isoformat()
    rows.append(mk(td, "W004", "Karthik R", "Changi P1", True, "08:00", "", 0, 0, False))
    return rows


# --------------------------------------------------------------- shape
def month_days(dfrom, dto):
    d1 = dt.date.fromisoformat(dfrom)
    d2 = dt.date.fromisoformat(dto)
    out, d = [], d1
    while d <= d2:
        out.append(d)
        d += dt.timedelta(days=1)
    return out


def worker_index(rows):
    workers = {}
    for a in rows:
        code = a.get("worker_code") or "?"
        workers.setdefault(code, a.get("worker") or "?")
    return dict(sorted(workers.items()))


def day_cell(a):
    """One day's value for the grid."""
    if a.get("present"):
        return round(float(a.get("nh", 0)) + float(a.get("ot", 0)), 1)
    return (a.get("absence") or "A").upper()[:2]


# --------------------------------------------------------------- excel
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(ws, cells_row):
    for c in cells_row:
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=RED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def build_excel(rows, days, today, path):
    wb = Workbook()

    # --- sheet 1: End-Time Verification (today) ---
    ws = wb.active
    ws.title = "End-Time Verification"
    ws.append([f"End-Time Verification  —  {today.strftime('%d/%m/%Y')}"])
    ws["A1"].font = Font(bold=True, size=13, color=RED)
    ws.append([])
    head = ["Code", "Name", "Site", "Present", "Start", "End", "NH", "OT", "Submitted", "Status"]
    ws.append(head)
    _hdr(ws, ws[3])
    trows = [a for a in rows if a.get("date") == today.isoformat()]
    trows.sort(key=lambda a: (a.get("site") or "", a.get("worker_code") or ""))
    pend_fill = PatternFill("solid", fgColor="FFC7CE")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    for a in trows:
        present = bool(a.get("present"))
        if not present:
            status = "Absent (" + (a.get("absence") or "absent") + ")"
        elif not a.get("end") or not a.get("submitted"):
            status = "PENDING"
        else:
            status = "Verified"
        ws.append([
            a.get("worker_code", ""), a.get("worker", ""), a.get("site", ""),
            "Yes" if present else "No", a.get("start", ""), a.get("end", ""),
            a.get("nh", 0) if present else "", a.get("ot", 0) if present else "",
            "Yes" if a.get("submitted") else "No", status,
        ])
        r = ws.max_row
        for col in range(1, 11):
            ws.cell(r, col).border = BORDER
            ws.cell(r, col).font = Font(size=9)
        if status == "PENDING":
            ws.cell(r, 10).fill = pend_fill
            ws.cell(r, 10).font = Font(bold=True, size=9)
        elif status == "Verified":
            ws.cell(r, 10).fill = ok_fill
    widths = [10, 22, 18, 9, 8, 8, 6, 6, 11, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # --- sheet 2: Monthly Attendance Sheet (grid) ---
    ws2 = wb.create_sheet("Monthly Attendance")
    ws2.append([f"Monthly Attendance Sheet  —  {today.strftime('%B %Y')}"])
    ws2["A1"].font = Font(bold=True, size=13, color=RED)
    ws2.append([])
    workers = worker_index(rows)
    cell_map = {}   # (code, date) -> row dict
    for a in rows:
        cell_map[(a.get("worker_code") or "?", a.get("date"))] = a
    head2 = ["Code", "Name"] + [d.strftime("%d") for d in days] + ["Days", "NH", "OT", "Total"]
    ws2.append(head2)
    _hdr(ws2, ws2[3])
    for code, name in workers.items():
        row_vals = [code, name]
        pdays = tnh = tot = 0.0
        for d in days:
            a = cell_map.get((code, d.isoformat()))
            if not a:
                row_vals.append("")
            else:
                row_vals.append(day_cell(a))
                if a.get("present"):
                    pdays += 1
                    tnh += float(a.get("nh", 0))
                    tot += float(a.get("ot", 0))
        row_vals += [int(pdays), round(tnh, 1), round(tot, 1), round(tnh + tot, 1)]
        ws2.append(row_vals)
        r = ws2.max_row
        for col in range(1, len(head2) + 1):
            ws2.cell(r, col).border = BORDER
            ws2.cell(r, col).font = Font(size=8)
            if col > 2:
                ws2.cell(r, col).alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 9
    ws2.column_dimensions["B"].width = 20
    for i in range(3, 3 + len(days)):
        ws2.column_dimensions[get_column_letter(i)].width = 4.5
    for i in range(3 + len(days), 3 + len(days) + 4):
        ws2.column_dimensions[get_column_letter(i)].width = 7
    ws2.freeze_panes = "C4"

    wb.save(path)
    log("wrote", path)


# ----------------------------------------------------------------- pdf
class PDF(FPDF):
    title = ""

    def header(self):
        self.set_font("Helvetica", "B", 13)
        self.set_text_color(0xC0, 0, 0)
        self.cell(0, 8, latin(self.title), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.set_draw_color(0xC0, 0, 0)
        self.set_line_width(0.5)
        y = self.get_y()
        self.line(self.l_margin, y, self.w - self.r_margin, y)
        self.ln(3)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(120, 120, 120)
        self.cell(0, 6, latin(f"VMMS - generated {dt.datetime.now(SGT).strftime('%d/%m/%Y %H:%M')} SGT   -   Page {self.page_no()}"), align="C")


def _pdf_table(pdf, headers, widths, rows, font=8, highlight=None):
    pdf.set_font("Helvetica", "B", font)
    pdf.set_fill_color(0xC0, 0, 0)
    pdf.set_text_color(255, 255, 255)
    for h, w in zip(headers, widths):
        pdf.cell(w, 6, latin(h), border=1, align="C", fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", font)
    for r in rows:
        fill = highlight(r) if highlight else False
        if fill:
            pdf.set_fill_color(255, 199, 206)
        for v, w in zip(r, widths):
            pdf.cell(w, 5.5, latin(v), border=1, align="C", fill=bool(fill))
        pdf.ln()


def build_pdf_verification(rows, today, path):
    pdf = PDF(orientation="P", unit="mm", format="A4")
    pdf.title = f"End-Time Verification  -  {today.strftime('%d/%m/%Y')}"
    pdf.set_auto_page_break(True, 15)
    pdf.add_page()
    trows = [a for a in rows if a.get("date") == today.isoformat()]
    trows.sort(key=lambda a: (a.get("site") or "", a.get("worker_code") or ""))
    headers = ["Code", "Name", "Site", "Pres", "Start", "End", "NH", "OT", "Subm", "Status"]
    widths = [16, 40, 34, 12, 16, 16, 12, 12, 14, 18]
    data, pend = [], 0
    for a in trows:
        present = bool(a.get("present"))
        if not present:
            status = "Absent"
        elif not a.get("end") or not a.get("submitted"):
            status = "PENDING"
            pend += 1
        else:
            status = "Verified"
        data.append([
            a.get("worker_code", ""), a.get("worker", ""), a.get("site", ""),
            "Y" if present else "N", a.get("start", ""), a.get("end", ""),
            a.get("nh", "") if present else "", a.get("ot", "") if present else "",
            "Y" if a.get("submitted") else "N", status,
        ])
    _pdf_table(pdf, headers, widths, data, font=8,
               highlight=lambda r: r[9] == "PENDING")
    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(0xC0, 0, 0)
    pdf.cell(0, 6, latin(f"{pend} worker(s) pending end-time verification."),
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.output(path)
    log("wrote", path)


def build_pdf_monthly(rows, days, today, path):
    pdf = PDF(orientation="L", unit="mm", format="A4")
    pdf.title = f"Monthly Attendance Sheet  -  {today.strftime('%B %Y')}"
    pdf.set_auto_page_break(True, 15)
    pdf.add_page()
    workers = worker_index(rows)
    cell_map = {}
    for a in rows:
        cell_map[(a.get("worker_code") or "?", a.get("date"))] = a

    name_w, tot_w = 34, 12
    avail = pdf.w - pdf.l_margin - pdf.r_margin - name_w - 3 * tot_w
    day_w = max(5.0, avail / max(len(days), 1))
    headers = ["Name"] + [d.strftime("%d") for d in days] + ["Days", "NH", "OT"]
    widths = [name_w] + [day_w] * len(days) + [tot_w, tot_w, tot_w]

    # header row
    pdf.set_font("Helvetica", "B", 6)
    pdf.set_fill_color(0xC0, 0, 0)
    pdf.set_text_color(255, 255, 255)
    for h, w in zip(headers, widths):
        pdf.cell(w, 6, latin(h), border=1, align="C", fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 6)
    for code, name in workers.items():
        pdays = tnh = tot = 0.0
        cells = [name if len(name) < 22 else name[:21]]
        for d in days:
            a = cell_map.get((code, d.isoformat()))
            if not a:
                cells.append("")
            else:
                cells.append(day_cell(a))
                if a.get("present"):
                    pdays += 1
                    tnh += float(a.get("nh", 0))
                    tot += float(a.get("ot", 0))
        cells += [int(pdays), round(tnh, 1), round(tot, 1)]
        for i, (v, w) in enumerate(zip(cells, widths)):
            pdf.cell(w, 5, latin(v), border=1, align="L" if i == 0 else "C")
        pdf.ln()
    pdf.output(path)
    log("wrote", path)


# ---------------------------------------------------------------- email
def send_email(subject, body, attachments):
    msg = EmailMessage()
    msg["From"] = envv("GMAIL_USER")
    msg["To"] = envv("REPORT_TO")
    msg["Subject"] = subject.replace("\n", " ").replace("\r", " ")
    msg.set_content(body)
    for path in attachments:
        with open(path, "rb") as f:
            data = f.read()
        sub = "pdf" if path.endswith(".pdf") else \
              "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        maintype = "application" if path.endswith((".pdf", ".xlsx")) else "application"
        msg.add_attachment(data, maintype=maintype, subtype=sub,
                           filename=os.path.basename(path))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=60) as s:
        s.login(envv("GMAIL_USER"), envv("GMAIL_APP_PASSWORD").replace(" ", ""))
        s.send_message(msg)
    log("email sent to", envv("REPORT_TO"))


# ----------------------------------------------------------------- main
def main():
    test = os.environ.get("REPORT_TEST") == "1"
    now = dt.datetime.now(SGT)
    today = now.date()
    dfrom = today.replace(day=1).isoformat()
    dto = today.isoformat()
    days = month_days(dfrom, dto)

    if test:
        rows = sample_rows(today)
        log(f"TEST mode — {len(rows)} sample rows")
    else:
        token = get_token()
        rows = fetch_attendance(token, dfrom, dto)
        log(f"fetched {len(rows)} attendance rows for {dfrom}..{dto}")

    outdir = os.environ.get("OUT_DIR", ".")
    os.makedirs(outdir, exist_ok=True)
    mtag = today.strftime("%Y-%m")
    dtag = today.strftime("%Y-%m-%d")
    f_xlsx = os.path.join(outdir, f"VMMS_Reports_{mtag}.xlsx")
    f_vpdf = os.path.join(outdir, f"EndTimeVerification_{dtag}.pdf")
    f_mpdf = os.path.join(outdir, f"MonthlyAttendance_{mtag}.pdf")

    build_excel(rows, days, today, f_xlsx)
    build_pdf_verification(rows, today, f_vpdf)
    build_pdf_monthly(rows, days, today, f_mpdf)

    pending = sum(
        1 for a in rows
        if a.get("date") == dto and a.get("present") and (not a.get("end") or not a.get("submitted"))
    )
    body = (
        f"VMMS site reports for {today.strftime('%d/%m/%Y')} (SGT).\n\n"
        f"Attached:\n"
        f"  1. End-Time Verification (today)  — PDF\n"
        f"  2. Monthly Attendance Sheet ({today.strftime('%B %Y')}) — PDF\n"
        f"  3. Both reports — Excel (2 sheets)\n\n"
        f"End-time verification: {pending} worker(s) still PENDING for today.\n\n"
        f"— Automated VMMS report. Review before use; not a payroll-final document."
    )
    attachments = [f_vpdf, f_mpdf, f_xlsx]

    if test:
        log("TEST mode — files built, e-mail NOT sent.")
        for p in attachments:
            log("  ", p, os.path.getsize(p), "bytes")
        return
    send_email(f"VMMS Reports — {today.strftime('%d/%m/%Y')}", body, attachments)


if __name__ == "__main__":
    main()
